import csv
from io import BytesIO, StringIO

from django.db.models import Count, Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from rest_framework import permissions, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from xhtml2pdf import pisa

from .cache_service import ReportCacheService
from .models import Ticket
from .serializers import TicketSerializer


class IsSellerOrAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated)


class TicketViewSet(viewsets.ModelViewSet):
    queryset = Ticket.objects.select_related('zone', 'draw_type', 'user').all().order_by('-id')
    serializer_class = TicketSerializer
    permission_classes = [IsSellerOrAdmin]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['get'])
    def pdf(self, request, pk=None):
        ticket = self.get_object()
        html = render_to_string('tickets/ticket.html', {'ticket': ticket})
        result = BytesIO()
        pisa.CreatePDF(src=html, dest=result)
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="ticket-{ticket.id}.pdf"'
        return response

    @action(detail=False, methods=['post'])
    def preview(self, request):
        data = request.data
        zone = data.get('zone')
        draw_type = data.get('draw_type')
        items = data.get('items', [])
        if not zone or not draw_type or not items:
            return Response({'detail': 'Datos incompletos.'}, status=400)
        # Renderizar previo sin persistir
        context = {
            'ticket': {
                'id': 'PREVIEW',
                'created_at': '',
                'zone': type('X', (), {'name': ''})(),
                'draw_type': type('X', (), {'name': ''})(),
                'user': type('X', (), {'username': request.user.username if request.user.is_authenticated else 'anon'})(),
                'total_pieces': sum(int(i.get('pieces', 0)) for i in items),
                'items': type('Y', (), {'all': lambda self: [type('I', (), i)() for i in items]})(),
            }
        }
        html = render_to_string('tickets/ticket.html', context)
        result = BytesIO()
        pisa.CreatePDF(src=html, dest=result)
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="ticket-preview.pdf"'
        return response

    def _build_filters(self, request):
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        filters = {}
        if start:
            filters['created_at__date__gte'] = start
        if end:
            filters['created_at__date__lte'] = end
        zone = request.query_params.get('zone')
        draw_type = request.query_params.get('draw_type')
        user = request.query_params.get('user')
        if zone:
            filters['zone_id__in'] = [z for z in str(zone).split(',') if z]
        if draw_type:
            filters['draw_type_id__in'] = [d for d in str(draw_type).split(',') if d]
        if user:
            filters['user_id__in'] = [u for u in str(user).split(',') if u]
        return filters

    def _build_summary(self, request):
        group_by = request.query_params.get('group_by', 'zone')
        if group_by not in {'zone', 'draw_type', 'user'}:
            return None, Response({'detail': 'group_by inválido'}, status=400)
        filters = self._build_filters(request)
        qs = Ticket.objects.filter(**filters)
        if group_by == 'zone':
            values = ('zone__name',)
            label_field = 'zone__name'
        elif group_by == 'draw_type':
            values = ('draw_type__name',)
            label_field = 'draw_type__name'
        else:
            values = ('user__username',)
            label_field = 'user__username'
        data = (
            qs.values(*values)
            .annotate(total_tickets=Count('id'), total_pieces=Sum('total_pieces'))
            .order_by()
        )
        result = [
            {
                'group': row[label_field] or '',
                'total_tickets': row['total_tickets'] or 0,
                'total_pieces': row['total_pieces'] or 0,
            }
            for row in data
        ]
        # Paginación
        try:
            page = max(1, int(request.query_params.get('page', '1')))
        except ValueError:
            page = 1
        try:
            page_size = min(500, max(1, int(request.query_params.get('page_size', '50'))))
        except ValueError:
            page_size = 50
        total_items = len(result)
        total_pages = (total_items + page_size - 1) // page_size if page_size else 1
        start = (page - 1) * page_size
        end = start + page_size
        paged = result[start:end]

        totals = qs.aggregate(total_tickets=Count('id'), total_pieces=Sum('total_pieces'))
        totals = {k: totals.get(k) or 0 for k in ['total_tickets', 'total_pieces']}

        # Totales diarios opcionales
        include_daily = request.query_params.get('daily', '0') in {'1', 'true', 'True'}
        daily = None
        if include_daily:
            daily_qs = (
                qs.values('created_at__date')
                .annotate(total_tickets=Count('id'), total_pieces=Sum('total_pieces'))
                .order_by('created_at__date')
            )
            daily = [
                {
                    'date': str(row['created_at__date']),
                    'total_tickets': row['total_tickets'] or 0,
                    'total_pieces': row['total_pieces'] or 0,
                }
                for row in daily_qs
            ]

        payload = {
            'summary': paged,
            'totals': totals,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_items': total_items,
                'total_pages': total_pages,
            },
        }
        if daily is not None:
            payload['daily'] = daily
        return payload, None

    @action(detail=False, methods=['get'], url_path='reports/summary')
    def reports_summary(self, request):
        # Obtener parámetros de la request
        start_date = request.query_params.get('start')
        end_date = request.query_params.get('end')
        zone = request.query_params.get('zone')
        draw_type = request.query_params.get('draw_type')
        user = request.query_params.get('user')
        group_by = request.query_params.get('group_by', 'zone')
        page = max(1, int(request.query_params.get('page', '1')))
        page_size = min(500, max(1, int(request.query_params.get('page_size', '50'))))
        include_daily = request.query_params.get('daily', '0') in {'1', 'true', 'True'}
        force_refresh = request.query_params.get('refresh', '0') in {'1', 'true', 'True'}
        fmt = request.query_params.get('format', '').lower()
        
        # Obtener reporte con cache
        payload = ReportCacheService.get_summary_report(
            start_date=start_date,
            end_date=end_date,
            zone=zone,
            draw_type=draw_type,
            user=user,
            group_by=group_by,
            page=page,
            page_size=page_size,
            include_daily=include_daily,
            force_refresh=force_refresh
        )
        
        # Si se solicita exportación, generar archivo
        if fmt in ['csv', 'xlsx']:
            try:
                filename = 'reporte'
                if fmt == 'csv':
                    sio = StringIO()
                    writer = csv.writer(sio)
                    writer.writerow(['Grupo', 'Total Tickets', 'Total Pedazos'])
                    for row in payload['summary']:
                        writer.writerow([row['group'], row['total_tickets'], row['total_pieces']])
                    writer.writerow([])
                    writer.writerow(['Totales', payload['totals']['total_tickets'], payload['totals']['total_pieces']])
                    if payload.get('daily'):
                        writer.writerow([])
                        writer.writerow(['Fecha', 'Total Tickets', 'Total Pedazos'])
                        for row in payload['daily']:
                            writer.writerow([row['date'], row['total_tickets'], row['total_pieces']])
                    out = HttpResponse(sio.getvalue(), content_type='text/csv; charset=utf-8')
                    out['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
                    return out
                elif fmt == 'xlsx':
                    wb = Workbook()
                    ws = wb.active
                    ws.title = 'Resumen'
                    ws.append(['Grupo', 'Total Tickets', 'Total Pedazos'])
                    for row in payload['summary']:
                        ws.append([row['group'], row['total_tickets'], row['total_pieces']])
                    ws.append([])
                    ws.append(['Totales', payload['totals']['total_tickets'], payload['totals']['total_pieces']])
                    if payload.get('daily'):
                        ws2 = wb.create_sheet('Diario')
                        ws2.append(['Fecha', 'Total Tickets', 'Total Pedazos'])
                        for row in payload['daily']:
                            ws2.append([row['date'], row['total_tickets'], row['total_pieces']])
                    bio = BytesIO()
                    wb.save(bio)
                    out = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    out['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
                    return out
            except Exception as e:
                print(f"Error en exportación: {e}")
                return Response({'detail': f'Error generando archivo de exportación: {str(e)}'}, status=500)
        
        return Response(payload)

    @action(detail=False, methods=['get'], url_path='test-export')
    def test_export(self, request):
        print(f"Test export endpoint called with params: {request.query_params}")
        return Response({"message": "Test export endpoint working", "params": dict(request.query_params)}, status=200)
    
    @action(detail=False, methods=['get', 'post'], url_path='reports/export')
    def reports_export(self, request):
        print(f"Export endpoint called with params: {request.query_params}")
        return Response({"message": "Export endpoint working", "params": dict(request.query_params)}, status=200)
        fmt = request.query_params.get('format', 'csv').lower()
        filename = 'reporte'
        if fmt == 'csv':
            sio = StringIO()
            writer = csv.writer(sio)
            writer.writerow(['Grupo', 'Total Tickets', 'Total Pedazos'])
            for row in payload['summary']:
                writer.writerow([row['group'], row['total_tickets'], row['total_pieces']])
            writer.writerow([])
            writer.writerow(['Totales', payload['totals']['total_tickets'], payload['totals']['total_pieces']])
            if payload.get('daily'):
                writer.writerow([])
                writer.writerow(['Fecha', 'Total Tickets', 'Total Pedazos'])
                for row in payload['daily']:
                    writer.writerow([row['date'], row['total_tickets'], row['total_pieces']])
            out = HttpResponse(sio.getvalue(), content_type='text/csv; charset=utf-8')
            out['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            return out
        elif fmt == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Resumen'
            ws.append(['Grupo', 'Total Tickets', 'Total Pedazos'])
            for row in payload['summary']:
                ws.append([row['group'], row['total_tickets'], row['total_pieces']])
            ws.append([])
            ws.append(['Totales', payload['totals']['total_tickets'], payload['totals']['total_pieces']])
            if payload.get('daily'):
                ws2 = wb.create_sheet('Diario')
                ws2.append(['Fecha', 'Total Tickets', 'Total Pedazos'])
                for row in payload['daily']:
                    ws2.append([row['date'], row['total_tickets'], row['total_pieces']])
            bio = BytesIO()
            wb.save(bio)
            out = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            out['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return out
        return Response({'detail': 'Formato no soportado'}, status=400)

    @action(detail=False, methods=['get'], url_path='cache/stats')
    def cache_stats(self, request):
        """Obtiene estadísticas del cache de reportes"""
        stats = ReportCacheService.get_cache_stats()
        return Response(stats)

    @action(detail=False, methods=['post'], url_path='cache/clear')
    def clear_cache(self, request):
        """Limpia el cache de reportes"""
        result = ReportCacheService.clear_all_reports_cache()
        return Response(result)


